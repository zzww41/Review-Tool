'''
@File:      CntWriter.py
@Time:      5/21/2020
@Author:    Zhou Wei
@Version:   1.0
@Contact:   Wei.ZHOU10@cn.bosch.com
'''

import xml.etree.ElementTree as ETree

from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl import worksheet
import re   # regular expression

from Error import *
from Logger import *
import operator
from global_definition import *

'''---------Global Array------------
   Save the data read from CNT
   block name, value ,id, size etc.
   ---------------------------------'''
#the order cannot be changed, if any new items need to be added, just add them in the rear
#{Datablock: [Value, id, Delivery, Reset To Delivery, Reprog, size]}
data_array = {}

class CntWriter:
    
    def __init__(self, file_from, file_to, sheet):
        self.start_from_rowNo = 0;
        self.file_from = file_from
        self.file_to = file_to 
        self.sheetName = sheet
        self.wb = 0
        self.tree = 0
        self.row_amount_valid = 0
        self.root = self._loadCNT()
        #self.errorObj = Error()
        self.errors = []
        self.errors_status = []
        #logging.basicConfig(filename = 'CNT.log', level = logging.DEBUG)

    #def _initArray(self, length):
     #   self.length = length
        #for i in range(0, length):
         #   data_array.append({})

    '''def recordLog(self, message, printOnScreen = False):
        
        logging.debug(message)
        if printOnScreen == True:
            print(message)'''

    def _loadSheet(self):
        self.wb = load_workbook(self.file_from, keep_vba = True, data_only = True)
        sheet = self.wb[self.sheetName]
        return sheet

    def _loadCNT(self):
        self.tree = ETree.parse(self.file_to)
            
        root = self.tree.getroot()
        return root

    def _saveCNT(self):
        self.tree.write(self.file_to)

    def checkDataErrors(self, value_str, size = 0, sizeCheck = False):
        #print("checkDataErrors--------")
        hex_list = str(value_str).split(' ')
        hex_size = len(hex_list)
        for i in range(hex_size):
            if len(hex_list[i]) != 2:
                return CHECK_DATA_LEN_WRONG
            
            char = hex_list[i]
            if re.match(CHECK_DATA_PATTERN,char) == None:
                return CHECK_DATA_CHAR_WRONG
            
        if sizeCheck == True:
            if hex_size != int(size):
                return CHECK_DATA_SIZE_WRONG

        
        return CHECK_DATA_CORRECT
                
        
    def _setValueForTemplate(self, array_dict, key, template):
        
        #if errors happened for input value read from excel, record it.
        error = self.checkDataErrors(array_dict[key][0], size = array_dict[key][5], sizeCheck = True) 
        if not operator.eq(error, CHECK_DATA_CORRECT):
            Error.appendErrors(key,error)
            
        #it's used for getting Hex value(e.g. 'FF 01 AB'), if the value is not hex but decimal, ascii, etc,
        #Converting the format to HEX is necessary, since only HEX is supported in this tool
        #value (array_dict[key][0])) could be int read from excel, so conver it to string first
        value_list = str(array_dict[key][0]).split(' ')
        data = ''
        list_len = len(value_list)
        for i in range(list_len):
            data += data_template.format(value_list[i])

        return template.format(key, data, list_len)

    #def createDataBlock():
    #    e1 = ETree.Element("DATABLOCK")
     ##   e1_sub = ETree.SubElement(e1, "DATABLOCK-NAME")
     #   DATAFORMAT


    def getEPK(self):
        data_block_list = self.root.find('MEM').findall(DATA_BLOCK)
        epk = ''
        for block in data_block_list:
            name = block.find(DATA_BLOCK_NAME).text
            if operator.eq(name, '__EEPROM_METADATA__'):
                data_list = block.findall('DATA')
                epk = data_list[1].text
                epk = epk[4:]
                print(epk)
        return epk

    def getSWVersion(self):
        sheet = self._loadSheet()
        return sheet[EXCEL_SW_VERSION].value

    def compareUUID(self):
        epk1 = str(self.getEPK())
        epk2 = str(self.getSWVersion())
        if epk1 == epk2:
            return True
        else:
            return False    
        
    
    def read(self):
        
        read_result = True
        Logger.recordLog("****Reading value from EXCEL*****\r\n")
        sheet = self._loadSheet()
        if not self.compareUUID():
            Error.appendErrors("EPK","EPK in EXCEL is not the same as that in CNT file")
            read_result = False
        #exclude header in the excel
        self.row_amount_valid = sheet.max_row + 1
        #this array can't be initialized in __init__ function, because valid row number is not known
        #self._initArray(row_amount_valid)
        for i in range(start_from_rowNo, self.row_amount_valid):
            cell_name_1 = data_mapping_to_cnt_dict[DATA_BLOCK_NAME] + str(i)
            isCellempty = sheet[cell_name_1].value
            #print(isCellempty)
            if operator.eq(isCellempty, None) or operator.eq(isCellempty, ""):
                continue
                #column No. + row No.  e.g 'A' + '2' = 'A2'
                
            cell_name_2 = data_mapping_to_cnt_dict[DATA_VALUE] + str(i)
            cell_name_3 = data_mapping_to_cnt_dict[DATA_ID] + str(i)
            cell_name_4 = data_mapping_to_cnt_dict[SESSION_DELIVERY] + str(i)
            cell_name_5 = data_mapping_to_cnt_dict[SESSION_RESET_TO_DELIVERY] + str(i)
            cell_name_6 = data_mapping_to_cnt_dict[SESSION_REPROG] + str(i)
            cell_name_7 = data_mapping_to_cnt_dict[DATA_SIZE] + str(i)    
            data_array[sheet[cell_name_1].value] = [sheet[cell_name_2].value,sheet[cell_name_3].value,sheet[cell_name_4].value,sheet[cell_name_5].value, sheet[cell_name_6].value, sheet[cell_name_7].value]
            #print(cell_name_3)
            i += 1
        Logger.recordLog(data_array)
        Logger.recordLog("\r\n*****End of Reading*****\r\n")

        return read_result

    def _elementNameList(self, element_list, element_type):
        element_text = []
        for i in range(0,len(element_list)):
            element_text.append(element_list[i].find(element_type).text)
        return element_text

    def modify_data_format(self):
        Logger.recordLog("****Modifying data format*****\r\n")
        #1st step: modify data format, e.g.<DATAFORMAT-IDENTIFIER>ee_datablock</DATAFORMAT-IDENTIFIER>
        session_list = self.root.find('MEM').find('SESSIONS').findall('SESSION')
        data_pointer_list = session_list[0].find('DATAPOINTERS').findall('DATAPOINTER')
        for i in range(0, len(data_pointer_list)):
            data_name = data_pointer_list[i].find('DATAPOINTER-NAME')
            if data_name.text not in data_array:
                continue
            #print(data_name.text)
            value = data_array[data_name.text][0]
            try:
                if not operator.eq(value,None) and not operator.eq(value,''):
                    data_pointer_list[i].find('DATAFORMAT-IDENTIFIER').text = 'ee_datablock'
                    Logger.recordLog("Changing data {0} to EE_DATABLOCK".format(data_name.text))
                else:
                    Logger.recordLog("The value of  {0} is empty, no need to set data format".format(data_name.text))
                    pass
            except NameError:
                Logger.recordLog("Name Error")
                pass
                
        #self._saveCNT()
        Logger.recordLog("****End of Modifying*****\r\n")
        
    def add_data_block(self):
        #2nd step: add a data block with Org as suffix, and put value in it
        Logger.recordLog("****adding data block*****\r\n")
        parent = self.root.find('MEM')
        datablock_list = parent.findall(DATA_BLOCK)
        datablock_namelist = self._elementNameList(datablock_list, DATA_BLOCK_NAME)
        for key in data_array:
            block_name = key
            string = block_name + TEXT_VALUE_EXSIT
            #if current value of the datablock has been added in cnt (if no data need to be added in the cnt, don't add the corresponding data block)
            if string in datablock_namelist:
                continue
            elif operator.eq(data_array[key][0], None) or operator.eq(data_array[key][0], '') or operator.eq(data_array[key][0], 'EMPTY'):
                continue
            #else
            else:
                new_blocks = self._setValueForTemplate(data_array, block_name, data_block_template)
                text = block_name + SEARCH_BY_TEXT
                idx = datablock_namelist.index(text) + 1
                new_elements = ETree.fromstringlist(new_blocks)
                element_list = new_elements.findall(DATA_BLOCK)
                for j in range(0, 1):
                    parent.insert(idx + j + 1,new_elements[j])
                    
        #self._saveCNT()
        Logger.recordLog("****End of adding data block*****\r\n")
        
    def setUseCases(self):
        Logger.recordLog("****Setting Use Cases*****\r\n")
        #if use case is already set in CNT file, no modification needed, because even if the previous value is not correct, eeedit will correct it after saving it.
        #so don't need to consider the case of update, only the case of adding '<></>' inside this use case is enough
        parent = self.root.find('MEM')
        sessions = parent.find('SESSIONS')
        session_list = sessions.findall('SESSION')
        idx = 0
        type_usecase = -1
        for j in range(2,5):
            datapointer_namelist = {}
            for key, value in data_array.items():
                value_str = str(value[j])
                if(operator.eq(value_str, 'Y')):
                    if j == 2:
                        type_usecase = USE_CASE_DELIVERY
                    elif j == 3:
                        type_usecase = USE_CASE_RESET
                    elif j == 4:
                        type_usecase = USE_CASE_REPROG
                    else:
                        continue
                elif operator.eq(value_str, 'N') or operator.eq(value_str, '') or value_str == 'None' or value_str == None:
                    continue
                else:
                    Error.appendErrors(key,value_str)
                    continue
                data_pointers_parent = 0
                datapointer_list = 0
                isUseCaseDefined = False  # used to check if some use case is in the xml file, if not, create a new session for that use case
                for idx in range(0,len(session_list)):
                    if USE_CASE_MAPPING[type_usecase] == session_list[idx].find('SESSION-NAME').text:
                        isUseCaseDefined = True
                        break
                if isUseCaseDefined == False:
                    #datapointer_text = data_pointer_template.format(key, value[1], key, 'ee_erase')
                    fail_mem_template = ETree.fromstring(session_FAILURE_MEMORY_template)
                    rpg_template = ETree.fromstring(session_REPROG_template)
                    sessions.append(fail_mem_template)
                    sessions.append(rpg_template)
                    session_list = sessions.findall('SESSION')
                    idx += 1   # append a new session(Reprog), so index would be increased by 1
                    
                for idx in range(0,len(session_list)):
                    if USE_CASE_MAPPING[type_usecase] == session_list[idx].find('SESSION-NAME').text:
                        break    
                data_pointers_parent = session_list[idx].find('DATAPOINTERS')
                datapointer_list = data_pointers_parent.findall('DATAPOINTER')
                datapointer_namelist = self._elementNameList(datapointer_list, DATA_POINTER_NAME)
                
                #if datablockname read from excel, is already set in xml file(CNT), don't add a new datapointer again
                #ignore the current datablockname, check the next one
                if key in datapointer_namelist:
                    continue
                
                string = ''
                if type_usecase == USE_CASE_DELIVERY:
                    if not operator.eq(value[0], '') and not operator.eq(value[0], None):
                        string = data_pointer_template.format(key, value[1], key, 'ee_datablock')
                        Logger.recordLog("----------DELIVERY-----------:")
                        Logger.recordLog(string)
                    else:
                        string = data_pointer_template.format(key, value[1], key, 'ee_range')
                elif type_usecase == USE_CASE_RESET:
                    string = data_pointer_template.format(key, value[1], key, 'ee_erase')
                    Logger.recordLog("----------USE_CASE_RESET-----------:", True)
                    Logger.recordLog(string, True)
                elif type_usecase == USE_CASE_REPROG:
                    string = data_pointer_template.format(key, value[1], key, 'ee_erase')
                    Logger.recordLog("----------USE_CASE_REPROG-----------:")
                    Logger.recordLog(string)
                else:
                    continue
                element = ETree.fromstring(string)
                data_pointers_parent.append(element)
        #self._saveCNT()
        Logger.recordLog("****End of Setting Use Cases*****\r\n")
            
        

    #return a list containing 4 values used to be added in data_block_template
    
        

    def addNewBlock(self):
        pass

    
        

    #return true if writing to CNT successfully, return fail otherwise
    def write(self):
        
        #1st step: modify data format, e.g.<DATAFORMAT-IDENTIFIER>ee_datablock</DATAFORMAT-IDENTIFIER>

        #2nd step: add a data block with Org as suffix, and put value in it
        #e.g.<DATABLOCK-NAME>NvM_ConfigId__Org</DATABLOCK-NAME>
        #<DATABLOCK>
        #  <DATABLOCK-NAME>NvM_ConfigId__Org</DATABLOCK-NAME>
        #  <DATAFORMAT>decimal</DATAFORMAT>
        #  <DATABLOCK-CLASS>eeprom_data_original</DATABLOCK-CLASS>
        #  <DATA>0</DATA>
        #  <DATA>61</DATA>
        #</DATABLOCK>

        #3rd deliver /reset/ reprog

        self.modify_data_format()
        self.setUseCases()
        self.add_data_block()
        if not Error.isErrorExisted():
            self._saveCNT()
            Logger.recordLog("********CNT Writing Succeeded********")
            return True
        else:
            Logger.recordLog("********CNT Writing Failed********\r\n")
            Logger.recordLog("********ERROR Details********\r\n")
            Logger.recordLog(Error.getErrors())
        return False
        
